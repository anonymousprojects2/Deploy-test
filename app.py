from flask import Flask, render_template, request, jsonify, session, redirect, url_for
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
import firebase_admin
from firebase_admin import credentials, auth
import qrcode
from PIL import Image
import cv2
from pyzbar.pyzbar import decode
import openpyxl
import os
import threading
import time
from datetime import datetime, timedelta
import json
import jwt
import bcrypt
from functools import wraps
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Initialize Flask app
app = Flask(__name__)
app.secret_key = os.urandom(24)
CORS(app, supports_credentials=True)

# Database configuration
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///attendmax.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# Initialize Firebase Admin SDK
cred = credentials.Certificate("test-e6dc9-firebase-adminsdk-fbsvc-b39a5a305c.json")
try:
    firebase_admin.initialize_app(cred)
except ValueError:
    pass

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key')
JWT_RESET_SECRET = os.environ.get('JWT_RESET_SECRET', 'reset-secret-key')
JWT_EXPIRATION = 3600  # 1 hour

# Database Models
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    role = db.Column(db.String(20), nullable=False)
    ip_address = db.Column(db.String(45))
    reset_token = db.Column(db.String(500))
    reset_token_expiry = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime)

    def set_password(self, password):
        self.password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())

    def check_password(self, password):
        return bcrypt.checkpw(password.encode('utf-8'), self.password_hash)

# Create database tables
with app.app_context():
    db.create_all()

# Middleware for IP validation
def validate_ip(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'status': 'error', 'message': 'Not authenticated'}), 401

        user = User.query.get(session['user_id'])
        if not user:
            return jsonify({'status': 'error', 'message': 'User not found'}), 404

        current_ip = request.remote_addr
        
        # If IP is not set (first login), set it
        if not user.ip_address:
            user.ip_address = current_ip
            db.session.commit()
        # If IP doesn't match, check for reset token
        elif user.ip_address != current_ip:
            if not user.reset_token or user.reset_token_expiry < datetime.utcnow():
                return jsonify({
                    'status': 'error',
                    'message': 'IP address mismatch. Please request a reset token.',
                    'code': 'IP_MISMATCH'
                }), 403
            
            # Verify reset token
            try:
                jwt.decode(user.reset_token, JWT_RESET_SECRET, algorithms=['HS256'])
            except jwt.InvalidTokenError:
                return jsonify({
                    'status': 'error',
                    'message': 'Invalid or expired reset token',
                    'code': 'INVALID_RESET_TOKEN'
                }), 403

        return f(*args, **kwargs)
    return decorated_function

# Routes
@app.route('/')
def home():
    return render_template('index.html')

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/login')
def login_page():
    return render_template('login.html')

@app.route('/auth/login', methods=['POST'])
def auth_login():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'status': 'error', 'message': 'No data received'}), 400

        email = data.get('username')
        password = data.get('password')
        role = data.get('role')

        if not all([email, password, role]):
            return jsonify({'status': 'error', 'message': 'Missing required fields'}), 400

        # Verify with Firebase
        try:
            firebase_user = auth.get_user_by_email(email)
        except auth.UserNotFoundError:
            return jsonify({'status': 'error', 'message': 'Invalid credentials'}), 401

        # Check local database
        user = User.query.filter_by(email=email).first()
        current_ip = request.remote_addr

        if not user:
            # First time login, create user
            user = User(email=email, role=role)
            user.set_password(password)
            user.ip_address = current_ip
            db.session.add(user)
        else:
            # Verify password and IP
            if not user.check_password(password):
                return jsonify({'status': 'error', 'message': 'Invalid credentials'}), 401
            
            if user.ip_address and user.ip_address != current_ip:
                return jsonify({
                    'status': 'error',
                    'message': 'IP address mismatch. Please request a reset token.',
                    'code': 'IP_MISMATCH'
                }), 403

        # Update last login
        user.last_login = datetime.utcnow()
        db.session.commit()

        # Set session
        session['user_id'] = user.id
        session['role'] = role
        session['email'] = email

        return jsonify({'status': 'success', 'message': 'Login successful', 'role': role})

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/auth/request-reset', methods=['POST'])
def request_reset():
    try:
        data = request.get_json()
        email = data.get('email')

        if not email:
            return jsonify({'status': 'error', 'message': 'Email is required'}), 400

        user = User.query.filter_by(email=email).first()
        if not user:
            return jsonify({'status': 'error', 'message': 'User not found'}), 404

        # Generate reset token
        reset_token = jwt.encode(
            {
                'user_id': user.id,
                'exp': datetime.utcnow() + timedelta(hours=1)
            },
            JWT_RESET_SECRET,
            algorithm='HS256'
        )

        # Store reset token
        user.reset_token = reset_token
        user.reset_token_expiry = datetime.utcnow() + timedelta(hours=1)
        db.session.commit()

        # In a real application, send this token via email
        # For demo purposes, we'll return it directly
        return jsonify({
            'status': 'success',
            'message': 'Reset token generated',
            'reset_token': reset_token
        })

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/auth/reset-ip', methods=['POST'])
def reset_ip():
    try:
        data = request.get_json()
        reset_token = data.get('reset_token')
        email = data.get('email')

        if not all([reset_token, email]):
            return jsonify({'status': 'error', 'message': 'Missing required fields'}), 400

        user = User.query.filter_by(email=email).first()
        if not user:
            return jsonify({'status': 'error', 'message': 'User not found'}), 404

        # Verify reset token
        try:
            jwt.decode(reset_token, JWT_RESET_SECRET, algorithms=['HS256'])
        except jwt.InvalidTokenError:
            return jsonify({'status': 'error', 'message': 'Invalid or expired reset token'}), 403

        if reset_token != user.reset_token or user.reset_token_expiry < datetime.utcnow():
            return jsonify({'status': 'error', 'message': 'Invalid or expired reset token'}), 403

        # Update IP address
        user.ip_address = request.remote_addr
        user.reset_token = None
        user.reset_token_expiry = None
        db.session.commit()

        return jsonify({
            'status': 'success',
            'message': 'IP address updated successfully'
        })

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Protected routes with IP validation
@app.route('/admin/dashboard')
@validate_ip
def admin_dashboard():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login_page'))
    return render_template('admin_dashboard.html')

@app.route('/student/dashboard')
@validate_ip
def student_dashboard():
    if 'user_id' not in session or session.get('role') != 'student':
        return redirect(url_for('login_page'))
    return render_template('student_dashboard.html')

# Initialize Excel file
EXCEL_FILE = "attendance_data.xlsx"
if not os.path.exists(EXCEL_FILE):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Username", "Department", "Year", "Timestamp"])
    workbook.save(EXCEL_FILE)

# Store QR codes in memory
valid_qr_codes = set()
qr_data = {}

def generate_qr_codes(department, year):
    """Generate QR codes every 15 seconds"""
    while True:
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        qr_data = f"{department}_{year}_{timestamp}"
        
        # Generate QR code
        qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
        qr.add_data(qr_data)
        qr.make(fit=True)
        
        # Save QR code
        img = qr.make_image(fill="black", back_color="white")
        img_path = f"static/qr_codes/{department}_{year}_latest.png"
        img.save(img_path)
        
        # Update valid QR codes
        valid_qr_codes.add(qr_data)
        if len(valid_qr_codes) > 4:  # Keep only last 4 QR codes valid
            valid_qr_codes.remove(min(valid_qr_codes))
        
        time.sleep(15)  # Generate new QR code every 15 seconds

@app.route('/admin/generate-qr', methods=['POST'])
@validate_ip
def generate_qr():
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401

    data = request.get_json()
    department = data.get('department')
    year = data.get('year')

    if not all([department, year]):
        return jsonify({'status': 'error', 'message': 'Missing required fields'}), 400

    # Start QR generation thread if not already running
    thread_key = f"{department}_{year}"
    if thread_key not in qr_data:
        qr_thread = threading.Thread(
            target=generate_qr_codes,
            args=(department, year),
            daemon=True
        )
        qr_data[thread_key] = {
            'thread': qr_thread,
            'department': department,
            'year': year
        }
        qr_thread.start()

    return jsonify({
        'status': 'success',
        'message': 'QR code generation started',
        'qr_path': f"/static/qr_codes/{department}_{year}_latest.png"
    })

@app.route('/student/mark-attendance', methods=['POST'])
@validate_ip
def mark_attendance():
    if 'user_id' not in session or session.get('role') != 'student':
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401

    data = request.get_json()
    qr_code = data.get('qr_code')

    if not qr_code:
        return jsonify({'status': 'error', 'message': 'No QR code data provided'}), 400

    if qr_code not in valid_qr_codes:
        return jsonify({'status': 'error', 'message': 'Invalid or expired QR code'}), 400

    # Record attendance in Excel
    try:
        department, year, _ = qr_code.split('_')
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([session['email'], department, year, timestamp])
        workbook.save(EXCEL_FILE)

        return jsonify({
            'status': 'success',
            'message': 'Attendance marked successfully'
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Error marking attendance: {str(e)}'
        }), 500

@app.route('/admin/attendance-data')
@validate_ip
def get_attendance_data():
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401

    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append({
                'username': row[0],
                'department': row[1],
                'year': row[2],
                'timestamp': row[3]
            })
        return jsonify({
            'status': 'success',
            'data': data
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Error fetching attendance data: {str(e)}'
        }), 500

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('home'))

if __name__ == '__main__':
    # Create required directories
    os.makedirs('static/qr_codes', exist_ok=True)
    
    app.run(debug=True, port=5000)
